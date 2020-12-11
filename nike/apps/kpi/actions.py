from os.path import dirname, abspath
import openpyxl, os
from apps.kpi.models import *
from apps.utils import *
path = dirname(abspath(__file__))
source_path = os.path.join(path, "sources")

''' 读取Excel文件数据 '''
def get_info_from_excel2(input_file_dic, input_file_sheets, file_column_dic, input_file):
    xl = openpyxl.load_workbook(input_file, data_only=True)
    xlsx_sheet_names = xl.sheetnames
    for sheet_name in xlsx_sheet_names:
        if sheet_name != '总数据表':
            continue
        table = xl[sheet_name]
        # list [Sheet1, Sheet2]
        input_file_sheets.append(sheet_name)
        # dict {Sheet1: [[1][2][3]], Sheet2: [[1][2][3]]}
        input_file_dic[sheet_name] = []
        # dict {Sheet1: [[0]], Sheet2: [[0]]}
        file_column_dic[sheet_name] = []
        max_row = table.max_row
        max_column = table.max_column
        # 本函数中第一行不为空
        # print(input_file, max_column)
        colume_info = []
        # 处理sheet为空/只有第一行的情况
        if max_column == 0:
            pass
        else:
            for i in range(1, max_column + 1):
                cellvalue = str(table.cell(1,i).value)
                if cellvalue is None:
                    cellvalue = ""
                colume_info.append(cellvalue)
        file_column_dic[sheet_name] = colume_info
        if max_row <= 1:
            input_file_dic[sheet_name].append(colume_info)
        else:
            for i in range(2, max_row + 1):
                row_info = []
                none_num = 0
                for j in range(1, max_column + 1):
                    cellvalue = str(table.cell(i,j).value)
                    if cellvalue == 'None' or cellvalue.strip(' ') == '':
                        none_num += 1
                        cellvalue = ""
                    row_info.append(cellvalue)
                if none_num == max_column:
                    continue
                input_file_dic[sheet_name].append(row_info)
    return input_file_dic, input_file_sheets, file_column_dic
def get_info_from_excel(input_file_dic, input_file_sheets, file_column_dic, input_file):
    xl = openpyxl.load_workbook(input_file, data_only=True)
    xlsx_sheet_names = xl.sheetnames
    for sheet_name in xlsx_sheet_names:
        #if sheet_name != '总数据表':
            #continue
        table = xl[sheet_name]
        # list [Sheet1, Sheet2]
        input_file_sheets.append(sheet_name)
        # dict {Sheet1: [[1][2][3]], Sheet2: [[1][2][3]]}
        input_file_dic[sheet_name] = []
        # dict {Sheet1: [[0]], Sheet2: [[0]]}
        file_column_dic[sheet_name] = []
        max_row = table.max_row
        max_column = table.max_column
        # 本函数中第一行不为空
        # print(input_file, max_column)
        colume_info = []
        # 处理sheet为空/只有第一行的情况
        if max_column == 0:
            pass
        else:
            for i in range(1, max_column + 1):
                cellvalue = str(table.cell(1,i).value)
                if cellvalue is None:
                    cellvalue = ""
                colume_info.append(cellvalue)
        file_column_dic[sheet_name] = colume_info
        if max_row <= 1:
            input_file_dic[sheet_name].append(colume_info)
        else:
            for i in range(2, max_row + 1):
                row_info = []
                none_num = 0
                for j in range(1, max_column + 1):
                    cellvalue = str(table.cell(i,j).value)
                    if cellvalue == 'None' or cellvalue.strip(' ') == '':
                        none_num += 1
                        cellvalue = ""
                    row_info.append(cellvalue)
                if none_num == max_column:
                    continue
                input_file_dic[sheet_name].append(row_info)
    return input_file_dic, input_file_sheets, file_column_dic
def get_info_from_excel1(input_file_dic, input_file_sheets, file_column_dic, input_file):
    xl = openpyxl.load_workbook(input_file, data_only=True)
    xlsx_sheet_names = xl.sheetnames
    for sheet_name in xlsx_sheet_names:
        table = xl[sheet_name]
        # list [Sheet1, Sheet2]
        input_file_sheets.append(sheet_name)
        # dict {Sheet1: [[1][2][3]], Sheet2: [[1][2][3]]}
        input_file_dic[sheet_name] = []
        # dict {Sheet1: [[0]], Sheet2: [[0]]}
        file_column_dic[sheet_name] = []
        max_row = table.max_row
        max_column = table.max_column
        # 本函数中第一行不为空
        # print(input_file, max_column)
        colume_info = []
        # 处理sheet为空/只有第一行的情况
        if max_column == 0:
            pass
        else:
            for i in range(1, max_column + 1):
                cellvalue = str(table.cell(1,i).value)
                if cellvalue is None:
                    cellvalue = ""
                colume_info.append(cellvalue)
        file_column_dic[sheet_name] = colume_info
        if max_row <= 1:
            input_file_dic[sheet_name].append(colume_info)
        else:
            for i in range(2, max_row + 1):
                row_info = []
                none_num = 0
                for j in range(1, max_column + 1):
                    cellvalue = str(table.cell(i,j).value)
                    if cellvalue == 'None' or cellvalue.strip(' ') == '':
                        none_num += 1
                        cellvalue = ""
                    row_info.append(cellvalue)
                if none_num == max_column:
                    continue
                input_file_dic[sheet_name].append(row_info)
    return input_file_dic, input_file_sheets, file_column_dic

def CenterKPI_to_DB(input_file):
    input_file_dic, input_file_sheets, file_column_dic = {}, [], {}
    input_file_dic, input_file_sheets, file_column_dic = get_info_from_excel2(input_file_dic,
                                                            input_file_sheets, file_column_dic, input_file)
    datas = input_file_dic[input_file_sheets[0]]
    result = {}
    csc_kpi = CSCKPI()
    task = csc_kpi.query.filter_by().all()
    if task:
        for res in task:
            db.session.delete(res)
            db.session.commit()
    center_kpi = CenterKPI()
    task = center_kpi.query.filter_by().all()
    if task:
        for res in task:
            db.session.delete(res)
            db.session.commit()
    index = 0
    input_dic, input_data = {}, []
    dc_list = []
    for i in range(len(datas)):
        if "大仓" in datas[i]:
            for ele in datas[i]:
                if (ele == '') or (not ele) or (ele == "大仓"):
                    continue
                if ele in dc_list:
                    continue
                dc_list.append(ele)
    dc_num = len(dc_list)
    print(dc_list)
    data_part1 = []
    for i in range(len(datas)):
        if (i <= 1):
            continue
        if "事故：件" in datas[i]:
            for j, dc in enumerate(dc_list):
                data_part1.append(['center', '当月', dc, datas[i][j+1], datas[i+1][j+1], datas[i+3][j+1], datas[i+4][j+1]])
                data_part1.append(['center', 'YTD', dc, datas[i][dc_num+1], datas[i+1][dc_num+1], datas[i+3][dc_num+1], datas[i+4][dc_num+1]])  
        if ('WK' in datas[i][0]) or ('月' in datas[i][0]) or ('YTD' in datas[i][0]):
            if datas[i][0] not in input_dic.keys():
                input_dic[datas[i][0]] = []
                for j, dc in enumerate(dc_list):
                    input_dic[datas[i][0]].append(['center', datas[i][0], dc, datas[i][j+1], datas[i][j+1+dc_num],
                                                    datas[i][j+1+2*dc_num], datas[i][j+1+3*dc_num], "", "", "", ""])
        if '库存毛差异：周' in datas[i]:
            index = i
            break
    db.session.execute(csc_kpi.__table__.insert(),
        [{
            "month": data[1], "city": data[2], "accident": data[3], "complain": data[4],
            "business_area": data[5], "usable_area": data[6]
        } for data in data_part1]
    )
    db.session.commit()
    for i in range(index, len(datas)):
        if ('WK' in datas[i][0]) or ('月' in datas[i][0]) or ('YTD' in datas[i][0]):
            if datas[i][0] in input_dic.keys():
                for data in input_dic[datas[i][0]]:
                    for j, dc in enumerate(dc_list):
                        if data[2] == dc:
                            data[7] = datas[i][j+1]
                            data[8] = datas[i][j+1+dc_num]
                            data[9] = datas[i][j+1+2*dc_num]
                            data[10] = datas[i][j+1+3*dc_num]
    for key in input_dic.keys():
        for data in input_dic[key]:
            input_data.append(data)
    db.session.execute(center_kpi.__table__.insert(),
        [{
            "week": data[1], "city": data[2], "price": data[3], "turnover": data[4],
            "B2C_efficiency": data[5], "B2B_efficiency": data[6], "stock": data[7],
            "performance": data[8], "profit": data[9], "score": data[10]
        } for data in input_data]
    )
    db.session.commit()

def isfloat(str):
    try:
        float(str)
        return True
    except ValueError:
        return False
from sqlalchemy.orm import class_mapper
def kpi_show():
    csc_kpi = CSCKPI()
    center_kpi = CenterKPI()
    [month_mark, week_mark] = get_target_month_week()
    same_task = csc_kpi.query.filter_by(month = "当月").all()
    ytd_task = csc_kpi.query.filter_by(month = "YTD").all()
    result = {}
    if (not same_task) or (not ytd_task):
        return result
    # 转换为数组
    same_datas, ytd_datas = [], []
    for same in same_task:
        same_datas.append([getattr(same, c.key) for c in class_mapper(same.__class__).columns])
    for ytd in ytd_task:
        ytd_datas.append([getattr(ytd, c.key) for c in class_mapper(ytd.__class__).columns])
    project1 = ['事故', '客户投诉', '可拓展新业务面积', '搭平台后可用面积']
    for i, ele in enumerate(project1):
        returnDatas = {
            'timeString': month_mark,
            'data': []
        }
        for row in same_datas:
            returnData = {
                'noid': '',
                'name': '',
                'ytd': 0,
                'num': 0,
                'target': 0
            }
            returnData['name'] = row[2]
            if isfloat(row[i + 3]):
                returnData['num'] = int(float(row[i + 3]))
            returnDatas['data'].append(returnData)
        for row in ytd_datas:
            for rd in returnDatas['data']:
                if rd['name'] != row[2]:
                    continue
                if isfloat(row[i + 3]):
                    rd['ytd'] = int(float(row[i + 3]))
        result[project1[i]] = returnDatas
    month_task = center_kpi.query.filter_by(week = month_mark).all()
    week_task = center_kpi.query.filter_by(week = week_mark).all()
    ytd_task = center_kpi.query.filter_by(week = "YTD").all()
    if (not month_task) or (not week_task) or (not ytd_task):
        return result
    month_datas, week_datas, ytd1_datas = [], [], []
    for ele in month_task:
        month_datas.append([getattr(ele, c.key) for c in class_mapper(ele.__class__).columns])
    for ele in week_task:
        week_datas.append([getattr(ele, c.key) for c in class_mapper(ele.__class__).columns])
    for ele in ytd_task:
        ytd1_datas.append([getattr(ele, c.key) for c in class_mapper(ele.__class__).columns])
    project2 = ['单件成本', '人员流失率', 'B2C人均效率', 'B2B人均效率', '库存净差异', '业绩达成率', '利润率']
    for i, ele in enumerate(project2):
        returnDatas = {
            'timeString': '',
            'data': []
        }
        if i in [2, 3, 4]:
            returnDatas['timeString'] = week_mark
            for row in week_datas:
                returnData = {
                    'noid': '',
                    'name': '',
                    'ytd': 0,
                    'num': 0,
                    'target': 0
                }
                returnData['name'] = row[2]
                if isfloat(row[i + 3]):
                    if i != 4:
                        returnData['num'] = round(float(row[i + 3]), 3)
                    else:
                        returnData['num'] = round(float(row[i + 3]) * 100, 3)
                returnDatas['data'].append(returnData)
            for row in ytd1_datas:
                for rd in returnDatas['data']:
                    if rd['name'] != row[2]:
                        continue
                    if isfloat(row[i + 3]):
                        if i != 4:
                            rd['ytd'] = round(float(row[i + 3]), 3)
                        else:
                            rd['ytd'] = round(float(row[i + 3]) * 100, 3)
        else:
            returnDatas['timeString'] = month_mark
            for row in month_datas:
                returnData = {
                    'noid': '',
                    'name': '',
                    'ytd': 0,
                    'num': 0,
                    'target': 0
                }
                returnData['name'] = row[2]
                if isfloat(row[i + 3]):
                    if i == 0:
                        returnData['num'] = round(float(row[i + 3]), 3)
                    else:
                        returnData['num'] = round(float(row[i + 3]) * 100, 3)
                returnDatas['data'].append(returnData)
            #print(returnDatas)
            for row in ytd1_datas:
                for rd in returnDatas['data']:
                    if rd['name'] != row[2]:
                        continue
                    if isfloat(row[i + 3]):
                        if i == 0:
                            rd['ytd'] = round(float(row[i + 3]), 3)
                        else:
                            rd['ytd'] = round(float(row[i + 3]) * 100, 3)
        result[project2[i]] = returnDatas
    return result

def kpi_merits(month):
    result = []
    input_file = os.path.join(source_path, '数据.xlsx')
    input_file_dic, input_file_sheets, file_column_dic = {}, [], {}
    input_file_dic, input_file_sheets, file_column_dic = get_info_from_excel1(input_file_dic,
                                                            input_file_sheets, file_column_dic, input_file)
    datas = input_file_dic["工作表" + month]
    data_num = len(datas)
    for i, data in enumerate(datas):
        returnData = {
            'rank': '',
            'name': '',
            'score': '',
            'color': 'red'
        }
        rank = str(data_num - i)
        returnData['rank'] = str(data_num - i)
        returnData['name'] = data[0]
        returnData['score'] = data[1]
        if i < 5:
            returnData['color'] ='green'
        result.append(returnData)
    # print(result)
    return result
def rank_show():
    [month_mark, week_mark] = get_target_month_week()
    center_kpi = CenterKPI()
    tasks = center_kpi.query.filter_by(week=week_mark).all()
    returnDatas = []
    color = ["red", "orange", "Cyan", "blue", "green", "white", "gray", "purple", "pink", "brown", "black", "pansy"]
    week_datas = []
    for ele in tasks:
        week_datas.append([getattr(ele, c.key) for c in class_mapper(ele.__class__).columns])
    for i, row in enumerate(week_datas):
        returnData = {
            'city': row[2],
            'color': color[i % len(color)],
            'score': row[10]
        }
        returnDatas.append(returnData)
    return [returnDatas]
def TransportKPI_to_DB(input_file):
    input_file_dic, input_file_sheets, file_column_dic = {}, [], {}
    input_file_dic, input_file_sheets, file_column_dic = get_info_from_excel(input_file_dic, input_file_sheets, file_column_dic, input_file)
    trans_kpi = TransportKPI()
    task = trans_kpi.query.filter_by().all()
    if task:
        for res in task:
            db.session.delete(res)
            db.session.commit()
    input_data = []
    for sheet in input_file_sheets:
        if (sheet == "总体说明") or sheet == "图表":
            continue
        datas = input_file_dic[sheet]
        temp1,temp2,temp3,temp4,temp5 = [sheet, '一月'],[sheet, '二月'],[sheet, '三月'],[sheet, '四月'],[sheet, '五月']
        temp6,temp7,temp8,temp9,temp10 = [sheet, '六月'],[sheet, '七月'],[sheet, '八月'],[sheet, '九月'],[sheet, '十月']
        temp11,temp12 = [sheet, '十一月'],[sheet, '十二月']
        for i in range(len(datas)):
            if i < 2:
                continue
            temp1.append(datas[i][3])
            temp2.append(datas[i][4])
            temp3.append(datas[i][5])
            temp4.append(datas[i][6])
            temp5.append(datas[i][7])
            temp6.append(datas[i][8])
            temp7.append(datas[i][9])
            temp8.append(datas[i][10])
            temp9.append(datas[i][11])
            temp10.append(datas[i][12])
            temp11.append(datas[i][13])
            temp12.append(datas[i][14])
        input_data += [temp1,temp2,temp3,temp4,temp5,temp6,temp7,temp8,temp9,temp10,temp11,temp12]
    db.session.execute(trans_kpi.__table__.insert(),
        [{
            "project": data[0], "month": data[1], "punctuality": data[2], "availability": data[3],
            "return_rate": data[4], "complaint": data[5], "accident": data[6], "collection_rate": data[7],
            "completion_rate": data[8], "profit_rate": data[9]
        } for data in input_data]
    )
    db.session.commit()
def show_month_transport():
    [month_mark, week_mark] = get_target_month_week()
    trans_kpi = TransportKPI()
    terms = ['交货准时率', '交货完好率', '回单返回率', '客诉次数', '安全事故次数', '回款率', '营业额完成率', '利润完成率', '表扬']
    returnDatas = {
        '交货准时率': {'timeString': month_mark, 'data': []},
        '交货完好率': {'timeString': month_mark, 'data': []},
        '回单返回率': {'timeString': month_mark, 'data': []},
        '客诉次数': {'timeString': month_mark, 'data': []},
        '安全事故次数': {'timeString': month_mark, 'data': []},
        '回款率': {'timeString': month_mark, 'data': []},
        '营业额完成率': {'timeString': month_mark, 'data': []},
        '利润完成率': {'timeString': month_mark, 'data': []},
        '表扬': {'timeString': month_mark, 'data': []},
        '异常分析': month_mark + '异常分析: 本月无异常。'
    }
    tasks = trans_kpi.query.filter_by(month=month_mark).all()
    month_datas, project_names = [], []
    for ele in tasks:
        month_datas.append([getattr(ele, c.key) for c in class_mapper(ele.__class__).columns])
    for row in month_datas:
        if row[1] not in project_names:
            project_names.append(row[1])
    for project in project_names:
        tasks = trans_kpi.query.filter_by(project=project).all()
        cur, ytd, cur1, ytd1, cur2, ytd2, cur3, ytd3 = 0, 0, 0, 0, 0, 0, 0, 0
        cur4, ytd4, cur5, ytd5, cur6, ytd6, cur7, ytd7 = 0, 0, 0, 0, 0, 0, 0, 0
        cur8, ytd8 = 0, 0
        for task in tasks:
            if task.punctuality:
                if task.month == month_mark:
                    cur = int(float(task.punctuality))
                ytd += int(float(task.punctuality))
            if task.availability:
                if task.month == month_mark:
                    cur1 = int(float(task.availability))
                ytd1 += int(float(task.availability))
            if task.return_rate:
                if task.month == month_mark:
                    cur2 = int(float(task.return_rate))
                ytd2 += int(float(task.return_rate))
            if task.complaint:
                if task.month == month_mark:
                    cur3 = int(float(task.complaint))
                ytd3 += int(float(task.complaint))
            if task.accident:
                if task.month == month_mark:
                    cur4 = int(float(task.accident))
                ytd4 += int(float(task.accident))
            if task.collection_rate:
                if task.month == month_mark:
                    cur5 = int(float(task.collection_rate))
                ytd5 += int(float(task.collection_rate))
            if task.completion_rate:
                if task.month == month_mark:
                    cur6 = int(float(task.completion_rate))
                ytd6 += int(float(task.completion_rate))
            if task.profit_rate:
                if task.month == month_mark:
                    cur7 = int(float(task.profit_rate))
                ytd7 += int(float(task.profit_rate))
        returnDatas['交货准时率']['data'].append({'project': project, 'cur': cur, 'ytd': ytd})
        returnDatas['交货完好率']['data'].append({'project': project, 'cur': cur1, 'ytd': ytd1})
        returnDatas['回单返回率']['data'].append({'project': project, 'cur': cur2, 'ytd': ytd2})
        returnDatas['客诉次数']['data'].append({'project': project, 'cur': cur3, 'ytd': ytd3})
        returnDatas['安全事故次数']['data'].append({'project': project, 'cur': cur4, 'ytd': ytd4})
        returnDatas['回款率']['data'].append({'project': project, 'cur': cur5, 'ytd': ytd5})
        returnDatas['营业额完成率']['data'].append({'project': project, 'cur': cur6, 'ytd': ytd6})
        returnDatas['利润完成率']['data'].append({'project': project, 'cur': cur7, 'ytd': ytd7})
        returnDatas['表扬']['data'].append({'project': project, 'cur': cur8, 'ytd': ytd8})
    return returnDatas
