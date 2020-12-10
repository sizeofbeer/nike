import os, re, time, sys
import openpyxl, datetime
from xlsxwriter.workbook import Workbook
from os.path import dirname, abspath

path = dirname(abspath(__file__))

parent_path = os.path.abspath(os.path.join(__file__, *(['..'] * 2)))
if parent_path not in sys.path:
    sys.path.append(parent_path)
tmp_path = os.path.join(parent_path, 'tmp')
if not os.path.exists(tmp_path):os.mkdir(tmp_path)

def format2stamp(date_time):
    tempstime = float(0)
    temp_date = str(date_time).replace("/", "-")
    if temp_date:
        try:
            tempstime = time.mktime(time.strptime(temp_date, "%Y-%m-%d"))
        except:
            try:
                tempstime = time.mktime(time.strptime(temp_date, "%Y-%m-%d %H:%M"))
            except:
                tempstime = time.mktime(time.strptime(temp_date, "%Y-%m-%d %H:%M:%S"))
    return tempstime

def stamp_format(date_time):
    date_str = ""
    temp_date = str(date_time).replace("/", "-")
    if temp_date:
        tempstime = float(0)
        try:
            tempstime = time.mktime(time.strptime(temp_date, "%Y-%m-%d"))
        except:
            try:
                tempstime = time.mktime(time.strptime(temp_date, "%Y-%m-%d %H:%M"))
            except:
                tempstime = time.mktime(time.strptime(temp_date, "%Y-%m-%d %H:%M:%S"))
        date_str = (datetime.datetime.utcfromtimestamp(tempstime)).strftime("%Y-%m-%d")
    return date_str

alpha_dic = {"A":0,"B":1,"C":2,"D":3,"E":4,"F":5,"G":6,"H":7,"I":8,"J":9,
            "K":10,"L":11,"M":12,"N":13,"O":14,"P":15,"Q":16,"R":17,"S":18,"T":19,
            "U":20,"V":21,"W":22,"X":23,"Y":24,"Z":25,"AA":26,"AB":27,"AC":28,"AD":29,
            "AE":30,"AF":31,"AG":32,"AH":33,"AI":34,"AJ":35,"AK":36,"AL":37,"AM":38,"AN":39}
map_list = ["A","B","C","D","E","F","G","H","I","J","K","L","M","N",
            "O","P","Q","R","S","T","U","V","W","X","Y","Z","AA","AB",
            "AC","AD","AE","AF","AG","AH","AI","AJ","AK","AL","AM","AN"]
def final_data_consist(origin_data, alpha_list):
    final_data = []
    for row in origin_data:
        temp = []
        for cell in alpha_list:
            if cell in alpha_dic.keys():
                cell_data = row[alpha_dic[cell]]
                temp.append(cell_data)
            else:
                temp.append(cell)
        final_data.append(temp)
    return final_data

def preChecking_inputFile(path, subject):
    file_list = []
    for _, _, folder in os.walk(path):
        for xls_file in folder:
            file_list.append(xls_file)
    for file_name in file_list:
        if re.search(subject + '(.*)',file_name, re.M|re.I):
            # print("载入文件：" + file_name)
            return file_name
    return None

def round_float(a, n):
    '''浮点数四舍五入功能'''
    w = 0.1
    for i in range(n + 2):
        w = w * 0.1
    a = a + w
    result = ""
    result_part = []
    int_part = int(a)
    result_part.append(int_part)
    a = a - int_part
    for i in range(n):
        a = a * 10
        int_part = int(a)
        a = a - int_part
        result_part.append(int_part)
    a = a * 10
    last_first_size = int(a)
    if last_first_size < 5:
        for i in range(len(result_part)):
            if i == 0:
                result = result + str(result_part[i]) + "."
            else:
                result = result + str(result_part[i])
    else:
        forward = 1
        for i in range(len(result_part)):
            if i == len(result_part)- 1:
                tag = result_part[len(result_part) - (1 + i)] + forward
                result = str(tag) + "." + result

            else:
                tag = result_part[len(result_part) -(1+i)] + forward
                if tag < 10:
                    forward = 0
                    result = str(tag) + result
                else:
                    result = str(tag%10) + result
    return result
''' 读入Excel文件数据 '''
def get_info_from_excel_v1(input_file_dic, input_file_sheets, file_column_dic, input_file):
    xl = openpyxl.load_workbook(input_file, data_only=True)
    xlsx_sheet_names = xl.sheetnames
    for sheet_name in xlsx_sheet_names:
        table = xl[sheet_name]
        # list [Sheet1, Sheet2]
        input_file_sheets.append(sheet_name)
        # dict {Sheet1: [[1][2][3]], Sheet2: [[1][2][3]]}
        input_file_dic[sheet_name] = []
        # dict {Sheet1: [[0]], Sheet2: [[0]]}
        file_column_dic[sheet_name] = []
        max_row = table.max_row
        max_column = table.max_column
        # 本函数中第一行不为空
        # print(input_file, max_column)
        colume_info = []
        # 处理sheet为空/只有第一行的情况
        if max_column == 0:
            pass
        else:
            for i in range(1, max_column + 1):
                cellvalue = table.cell(1,i).value
                if cellvalue is None:
                    cellvalue = ""
                colume_info.append(cellvalue)
        file_column_dic[sheet_name] = colume_info
        if max_row <= 1:
            input_file_dic[sheet_name].append(colume_info)
        else:
            for i in range(2, max_row + 1):
                row_info = []
                for j in range(1, max_column + 1):
                    cellvalue = table.cell(i,j).value
                    if cellvalue is None:
                        cellvalue = ""
                    row_info.append(cellvalue)
                input_file_dic[sheet_name].append(row_info)
    return input_file_dic, input_file_sheets, file_column_dic

from xlrd import xldate_as_tuple
import xlrd
''' 读取Excel文件 '''
def get_info_from_excel_v2(input_file_dic, input_file_sheets, file_column_dic, input_file):
    xl = xlrd.open_workbook(input_file)        # 打开Excel文件
    xlsx_sheet_names =xl.sheet_names()         # 获取工作表名称
    for sheet_name in xlsx_sheet_names:
        table = xl.sheet_by_name(sheet_name)   # 获取工作表对象
        # list [Sheet1, Sheet2]
        input_file_sheets.append(sheet_name)
        # dict {Sheet1: [[1][2][3]], Sheet2: [[1][2][3]]}
        input_file_dic[sheet_name] = []
        # dict {Sheet1: [[0]], Sheet2: [[0]]}
        file_column_dic[sheet_name] = []

        max_row = table.nrows
        max_column = table.ncols
        # 本函数中第一行不为空
        # print(input_file, max_column)
        colume_info = []
        # 处理sheet为空/只有第一行的情况        
        if max_column == 0:
            pass
        else:
            colume_info = table.row_values(0)[:max_column]
        file_column_dic[sheet_name] = colume_info
        if max_row <= 1:
            input_file_dic[sheet_name].append(colume_info)
        else:
            for i in range(1, max_row):
                row_info = []
                for j in range(0, max_column):
                    ctype = table.cell(i, j).ctype
                    cell = table.cell_value(i, j)
                    if ctype == 2:
                        if cell % 1 == 0: # ctype为2且为实际为整数
                            cell = int(cell)
                    if ctype == 3:
                        date = datetime.datetime(*xldate_as_tuple(cell, 0))
                        cell = date.strftime('%Y-%m-%d %H:%M:%S')
                    row_info.append(cell)
                input_file_dic[sheet_name].append(row_info)
    return input_file_dic, input_file_sheets, file_column_dic

''' 写入Excel文件 '''
def input_excel(xlsx_name, head_in, data_in, SheetName_in):
    workbook = Workbook(xlsx_name)
    for i in range(len(SheetName_in)):
        SheetName = SheetName_in[i]
        heading = head_in[i]
        data = data_in[i]
        worksheet = workbook.add_worksheet(SheetName)
        bold = workbook.add_format({'bold': 1, 'border': 1, 'align': 'center','valign': 'vcenter', "fg_color": "yellow"})
        bold1 = workbook.add_format({'border': 1,'align': 'center','valign': 'vcenter'})
        worksheet.write_row('A1', heading, bold)
        for row_num, row_data in enumerate(data):
            worksheet.write_row(row_num + 1, 0, row_data, bold1)
    workbook.close()

def differ_input_excel(xlsx_name, head_in, data_in, location_in, SheetName_in):
    # , SheetName, headings, data
    workbook = Workbook(xlsx_name)
    for i in range(len(SheetName_in)):
        # print(i, SheetName_in[i])
        SheetName = SheetName_in[i]
        headings = head_in[i]
        data = data_in[i]
        location = location_in[i]
        # print(location)
        # print(headings)
        worksheet = workbook.add_worksheet(SheetName)
        bold = workbook.add_format({'bold': 1, 'border': 1, 'align': 'center','valign': 'vcenter', "fg_color": "yellow"})
        bold1 = workbook.add_format({'border': 1,'align': 'center','valign': 'vcenter'})
        bold2 = workbook.add_format({'bold': 1, 'color':'red','border': 1,'align': 'center','valign': 'vcenter'})
        worksheet.write_row('A1', headings, bold)
        for row_num, row_data in enumerate(data):
            worksheet.write_row(row_num + 1, 0, row_data, bold1)
            if location:
                for local in location[int((row_num) / 2)]:
                    worksheet.write(map_list[int(local)] + str(row_num + 2), row_data[int(local)], bold2)
    workbook.close()
def get_the_month_and_week():
    # 获取当前日期
    time_now = datetime.datetime.now()
    # 提取当前日期的信息
    week_id = int(time.strftime("%W"))  # 当前日期所在的全年的星期ID，注：若1月1号为星期1，则1月1号的全年星期ID为1；否则，为0
    weekday = time_now.weekday()+1 # 星期几，1-星期一；2-星期二；。。。；7-星期日
    year = time_now.year # 年
    month = time_now.month #月
    day = time_now.day # 日
    print(week_id, weekday, year,month, day)
    # 若指标展示的是月指标，则展示上月指标，这里返回上月所属的年show_month_year和show_month_year
    if month == 1:
        show_month_year = year - 1
        show_month_month = 12
    else:
        show_month_year = year
        show_month_month = month -1

    # 若指标展示的是周指标且当前周为当前月的第一个周，即该周的星期一为第一个星期一
    # 判断条件为：1<= (day + 7 - weekday)/7 <2  是否成立。若成立，则为第一个周；
    #     此种情况下： 若 month = 1 则：show_week_year = year - 1, show_week_month_or_week = 12, tag_month_or_week = 1 ,注tag_month_or_week值：1表示月；0表示week
    #               否则，show_week_year = year, show_week_month_or_week = month - 1, tag_month_or_week = 1
    # 否则，有两种情况。
    #    第一种情况：(day + 7 - weekday)/7 < 1
    #       此种情况下： 若 month = 1 则：show_week_year = year - 1, show_week_month_or_week = 上一年倒数第二周, tag_month_or_week = 0 ,注tag_month_or_week值：1表示月；0表示week
    #   week = datetime.datetime.strptime(str(year-1) + str(month) + str(day),"%Y%m%d").strftime("%W")
    #                否则，show_week_year = year, show_week_month_or_week = week_id - 1, tag_month_or_week = 0
    #   第二种情况：(day + 7 - weekday)/7 = 0
    #           show_week_year = year, show_week_month_or_week = week_id - 1, tag_month_or_week = 0

    if_first_week = abs((day + 7 - weekday)/7)
    # print(if_first_week)

    if if_first_week < 1:
        tag_month_or_week = 0
        if month == 1:
            show_week_year = year - 1
            # if day < 10:
            #     day_str = "0" + str(day)
            # else:
            #     day_str = "0" + str(day)
            show_week_month_or_week = str(int(datetime.datetime.strptime(str(show_week_year) + "1231","%Y%m%d").strftime("%W")) - 1)
        else:
            show_week_year = year
            show_week_month_or_week = week_id - 1
    elif if_first_week < 2:
        tag_month_or_week = 1
        if month == 1:
            show_week_year = year - 1
            show_week_month_or_week = 12
        else:
            show_week_year = year
            show_week_month_or_week = month - 1
    else:
        tag_month_or_week = 0
        show_week_year = year
        show_week_month_or_week = week_id - 1

    return int(show_month_year), int(show_month_month), int(show_week_year), int(show_week_month_or_week), int(tag_month_or_week)
month_dic = {1: "一月",2: "二月",3: "三月",4: "四月",5: "五月",6: "六月",7: "七月",8: "八月",9: "九月",10: "十月",11: "十一月",12: "十二月"}
def get_target_month_week():
    show_month_year, show_month_month, show_week_year, show_week_month_or_week, tag_month_or_week = get_the_month_and_week()
    # print(show_month_year, show_month_month, show_week_year, show_week_month_or_week, tag_month_or_week)
    print("指标展示的是月，则应展示的月份为：", show_month_year, "年的", show_month_month, "月指标")
    month = month_dic[show_month_month]
    week = ''
    if tag_month_or_week ==0:
        print("指标展示的是周且不是所在月第1周，则应展示的周为：", show_week_year, "年的第", show_week_month_or_week, "周指标")
        week = "WK" + str(show_week_month_or_week)
    else:
        print("指标展示的是周且是所在月第1周，则应展示的上月：", show_week_year, "年的", show_week_month_or_week, "月指标")
        week = month_dic[show_week_month_or_week]
    return [month, week]
import smtplib
from email.header import Header
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
def send_email(xlsx_path):
    send_finish_tag = True
    ''' qq邮箱 2 企业邮箱 '''
    sender = 'jimmy.ma@runbow.com.cn'
    receiver = 'luke.lu@runbow.com.cn,vicky.li@runbow.com.cn,jimmy.ma@runbow.com.cn'
    smtpserver = 'smtp.qiye.163.com'
    username = 'jimmy.ma@runbow.com.cn'
    password = 'XYfWUybVxqKS1dTh'

    # 创建一个带附件的实例
    message = MIMEMultipart()
    message['From'] = sender
    message['To'] = receiver
    today = datetime.datetime.today().strftime('%Y-%m-%d')
    message['Subject'] = Header(today + "出货预报", 'utf-8')

    # 邮件正文内容
    message.attach(MIMEText("FYI", 'plain', 'utf-8'))
    # 构造附件1（附件为TXT格式的文本）
    att = MIMEText(open(xlsx_path, 'rb').read(), 'base64', 'utf-8')
    att["Content-Type"] = 'application/octet-stream'
    att["Content-Disposition"] = 'attachment; filename="Est_Report.xlsx"'
    message.attach(att)
    try:
    #if True:
        smtpObj = smtplib.SMTP_SSL(smtpserver)  # 注意：如果遇到发送失败的情况（提示远程主机拒接连接），这里要使用SMTP_SSL方法
        print("邮件远程主机连接成功")
        smtpObj.connect(smtpserver)
        print("邮件远程服务连接成功")
        smtpObj.login(username, password)
        print("邮件用户名密码登录成功")
        smtpObj.sendmail(sender, receiver, message.as_string())
        print("信息处理结果的邮件发送成功！！！")
        smtpObj.quit()
    except:
        print("处理结果邮件发送异常")
        send_finish_tag = False
    return send_finish_tag